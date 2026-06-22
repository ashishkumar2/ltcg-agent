from __future__ import annotations

from pathlib import Path

from ltcg_agent.models.tax import TaxSummary


class ReportGenerator:
    def generate(
        self,
        summary: TaxSummary,
        explanation: str,
        output_path: Path,
    ) -> Path:
        output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        _write_xlsx(summary, explanation, output_path)
        return output_path


def _write_xlsx(summary: TaxSummary, explanation: str, path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    _write_summary_sheet(wb.active, summary)
    _write_events_sheet(wb.create_sheet("Tax Events"), summary)
    _write_provenance_sheet(wb.create_sheet("FX Provenance"), summary)
    _write_schedule_fa_sheet(wb.create_sheet("Schedule FA"), summary)
    _write_harvest_sheet(wb.create_sheet("Harvest Candidates"), summary)
    _write_escalation_sheet(wb.create_sheet("Escalations"), summary)
    _write_explanation_sheet(wb.create_sheet("Explanation"), explanation)
    _write_disclaimer_sheet(wb.create_sheet("Disclaimer"), summary)
    wb.save(path)


def _write_summary_sheet(ws, summary: TaxSummary) -> None:
    rows = [
        ("Financial Year", summary.financial_year),
        ("Total LTCG (INR)", summary.total_ltcg.to_major_units()),
        ("Total STCG (INR)", summary.total_stcg.to_major_units()),
        ("Total LTCL (INR)", summary.total_ltcl.to_major_units()),
        ("Total STCL (INR)", summary.total_stcl.to_major_units()),
        ("Number of Disposal Events", len(summary.events)),
        ("Escalation Flags", len(summary.escalation_flags)),
        ("Must Escalate to CA", summary.must_escalate),
    ]
    for row in rows:
        ws.append(list(row))


def _write_events_sheet(ws, summary: TaxSummary) -> None:
    ws.append([
        "Ticker", "Lot ID", "Acquisition Date", "Disposal Date",
        "Quantity", "Cost Basis INR", "Sale Proceeds INR",
        "Gain/Loss INR", "Category", "Holding Days", "Grandfathered",
    ])
    for e in summary.events:
        ws.append([
            e.ticker,
            e.lot_id,
            e.acquisition_date.isoformat(),
            e.disposal_date.isoformat(),
            e.quantity,
            e.cost_basis_inr.to_major_units(),
            e.sale_proceeds_inr.to_major_units(),
            e.gain_inr.to_major_units(),
            e.category.value,
            e.holding_days,
            e.grandfathered,
        ])


def _write_provenance_sheet(ws, summary: TaxSummary) -> None:
    ws.append([
        "Ticker", "Lot ID", "Side",
        "USD Cents", "TTBR Paise/USD", "Rate Date", "INR Paise",
    ])
    for e in summary.events:
        for side, fx in (("Acquisition", e.acquisition_fx), ("Disposal", e.disposal_fx)):
            ws.append([
                e.ticker,
                e.lot_id,
                side,
                fx.usd_cents,
                fx.ttbr_paise_per_usd,
                fx.rate_date.isoformat(),
                fx.inr_paise,
            ])


def _write_schedule_fa_sheet(ws, summary: TaxSummary) -> None:
    ws.append(["Schedule FA — Foreign Assets (US Equities)"])
    ws.append(["Country", "Nature of Asset", "Date of Acquisition", "Cost (INR)"])
    seen: set[str] = set()
    for e in summary.events:
        key = f"{e.ticker}_{e.acquisition_date}"
        if key in seen:
            continue
        seen.add(key)
        ws.append([
            "United States",
            f"Listed Equity — {e.ticker}",
            e.acquisition_date.isoformat(),
            e.cost_basis_inr.to_major_units(),
        ])


def _write_harvest_sheet(ws, summary: TaxSummary) -> None:
    ws.append(["Ticker", "Lot ID", "Acquisition Date", "Qty", "Unrealised Loss INR", "Category"])
    for h in summary.harvest_candidates:
        ws.append([
            h.ticker,
            h.lot_id,
            h.acquisition_date.isoformat(),
            h.quantity,
            h.unrealised_loss_inr.to_major_units(),
            h.category.value,
        ])


def _write_escalation_sheet(ws, summary: TaxSummary) -> None:
    ws.append(["Severity", "Reason", "Detail"])
    for flag in summary.escalation_flags:
        ws.append([flag.severity.value, flag.reason.value, flag.detail])
    if not summary.escalation_flags:
        ws.append(["INFO", "NONE", "No escalation flags raised for this portfolio."])


def _write_explanation_sheet(ws, explanation: str) -> None:
    ws.append(["LLM Explanation (grounded in IT Act — not tax advice)"])
    ws.append([explanation])


def _write_disclaimer_sheet(ws, summary: TaxSummary) -> None:
    ws.append(["DISCLAIMER"])
    ws.append([summary.disclaimer])
